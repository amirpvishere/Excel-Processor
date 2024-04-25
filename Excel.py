import openpyxl as xl


def excel_processor(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        prices = sheet.cell(row=row, column=1).value
        discounts = sheet.cell(row=row, column=2).value
        final_prices = sheet.cell(row=row, column=3)
        final_price = prices * discounts
        final_prices.value = final_price

    wb.save(filename)


excel_processor("book1.xlsx")
